from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_ROOT = APP_ROOT.parent
OUTPUT_PATH = APP_ROOT / "data" / "generated" / "normalized.json"
PRODUCT_MASTER_RELATIVE_PATH = Path("sample-data/product-master.xlsx")
VAT_RATE = 0.19
PROVISION_RATE = 0.15


def repair_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    repaired = value.replace("\ufeff", "").strip()
    for _ in range(2):
        if not any(marker in repaired for marker in ("Ã", "â", "Â", "ð")):
            break
        try:
            candidate = repaired.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            break
        if candidate == repaired:
            break
        repaired = candidate
    return repaired


def sku_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return repair_text(str(value)).strip()


def clean_id(value: Any) -> str:
    text = sku_text(value)
    match = re.fullmatch(r'=\"(.*)\"', text)
    return match.group(1) if match else text


def number(value: Any, *, percent: bool = False, comma_decimal: bool = False) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        result = float(value)
        return result / 100 if percent and result > 1 else result
    text = repair_text(str(value))
    if not text or text in {"—", "-", "<5%"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if not text or text == "-":
        return None
    if "," in text and "." in text:
        if text.rfind(",") < text.rfind("."):
            text = text.replace(",", "")
        else:
            text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        tail = text.rsplit(",", 1)[1]
        if comma_decimal or len(tail) != 3:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    try:
        result = float(text)
    except ValueError:
        return None
    if negative:
        result = -result
    if percent:
        result /= 100
    return result


def metric(value: float | None) -> float:
    return 0.0 if value is None or math.isnan(value) else value


def ratio(numerator: float, denominator: float) -> float | None:
    return numerator / denominator if denominator else None


def round_value(value: float | None, digits: int = 4) -> float | None:
    return None if value is None else round(value, digits)


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = repair_text(str(value or ""))
    for pattern in ("%b %d, %Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def read_csv(file_path: Path) -> tuple[list[dict[str, str]], list[str]]:
    text = file_path.read_bytes().decode("utf-8-sig", errors="replace")
    sample = text[:65536]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    raw_rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    if not raw_rows:
        return [], []
    headers = [repair_text(header) for header in raw_rows[0]]
    rows = []
    for raw_row in raw_rows[1:]:
        padded = raw_row + [""] * max(0, len(headers) - len(raw_row))
        rows.append({headers[index]: repair_text(padded[index]) for index in range(len(headers))})
    return rows, headers


def read_xlsx(file_path: Path, sheet_index: int = 0) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[sheet_index]
    iterator = sheet.iter_rows(values_only=True)
    headers = [repair_text(value) for value in next(iterator)]
    rows = []
    for raw_row in iterator:
        if not any(value not in (None, "") for value in raw_row):
            continue
        rows.append({headers[index]: repair_text(raw_row[index]) for index in range(min(len(headers), len(raw_row)))})
    workbook.close()
    return rows


def hash_file(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def aggregate_metrics(rows: list[dict[str, Any]], mapping: dict[str, str]) -> dict[str, float]:
    totals = {target: 0.0 for target in mapping.values()}
    for row in rows:
        for source, target in mapping.items():
            totals[target] += metric(number(row.get(source)))
    return totals


def infer_category(margin: float | None) -> str | None:
    if margin is None:
        return None
    if margin < 0.20:
        return "Abverkauf"
    if margin <= 0.25:
        return "Preiseinstieg"
    if margin <= 0.30:
        return "Core"
    if margin <= 0.35:
        return "Performance"
    return "Premium"


def enrich_metrics(record: dict[str, Any]) -> dict[str, Any]:
    impressions = metric(record.get("impressions"))
    clicks = metric(record.get("clicks"))
    spend = metric(record.get("spend"))
    purchases = metric(record.get("purchases"))
    sales = metric(record.get("sales"))
    record.update({
        "ctr": round_value(ratio(clicks, impressions)),
        "cvr": round_value(ratio(purchases, clicks)),
        "acos": round_value(ratio(spend, sales)),
        "roas": round_value(ratio(sales, spend)),
        "cpc": round_value(ratio(spend, clicks)),
        "cpa": round_value(ratio(spend, purchases)),
        "aov": round_value(ratio(sales, purchases)),
    })
    for key in ("impressions", "clicks", "spend", "purchases", "sales", "units"):
        if key in record:
            record[key] = round_value(metric(record[key]), 2)
    return record


def default_product_master(source_root: Path) -> Path:
    return (source_root / PRODUCT_MASTER_RELATIVE_PATH).resolve()


def find_source_files(source_root: Path, product_master: Path | None = None) -> dict[str, Path]:
    expected = {
        "advertised_product": source_root / "sample-data" / "advertised-product.csv",
        "campaign": source_root / "sample-data" / "campaign.csv",
        "placement": source_root / "sample-data" / "placement.csv",
        "search_term_summary": source_root / "sample-data" / "search-term-summary.csv",
        "search_term_daily": source_root / "sample-data" / "search-term-daily.csv",
        "targeting": source_root / "sample-data" / "targeting.csv",
        "business_report": source_root / "sample-data" / "business-report.csv",
        "product_master": (product_master or default_product_master(source_root)).resolve(),
        "amazon_listing": source_root / "sample-data" / "amazon-product-list.xlsx",
        "economics": source_root / "sample-data" / "calculation-workbook.xlsx",
    }
    missing = [str(path) for path in expected.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing required source files:\n" + "\n".join(missing))
    return expected


def build_snapshot(source_root: Path, product_master: Path | None = None) -> dict[str, Any]:
    files = find_source_files(source_root, product_master)
    advertised_rows, _ = read_csv(files["advertised_product"])
    campaign_rows, _ = read_csv(files["campaign"])
    placement_rows, _ = read_csv(files["placement"])
    search_summary_rows, _ = read_csv(files["search_term_summary"])
    search_daily_rows, _ = read_csv(files["search_term_daily"])
    targeting_rows, _ = read_csv(files["targeting"])
    business_rows, _ = read_csv(files["business_report"])
    master_rows = read_xlsx(files["product_master"])
    product_rows = read_xlsx(files["amazon_listing"])
    economics_rows = read_xlsx(files["economics"])

    economics: dict[str, dict[str, Any]] = {}
    for row in economics_rows:
        sku = sku_text(row.get("SKU"))
        if not sku:
            continue
        margin = number(row.get("Margin (%)"), percent=False, comma_decimal=True)
        sales_net = number(row.get("VK (€ netto)"), comma_decimal=True)
        landed = number(row.get("Einstandspreis (€ netto)"), comma_decimal=True)
        if margin is None and sales_net and landed is not None:
            margin = (sales_net - landed) / sales_net
        category = repair_text(row.get("Kategorie")) or infer_category(margin)
        economics[sku] = {
            "margin": round_value(margin),
            "category": category,
            "description": repair_text(row.get("Bezeichnung")),
            "amazonAsin": clean_id(row.get("Amazon ASIN")),
            "purchaseCostNet": round_value(number(row.get("EK (€ netto)"), comma_decimal=True), 2),
            "deliveryCostNet": round_value(number(row.get("Logistic Cost (€ netto)"), comma_decimal=True), 2),
        }

    economics_asin_to_sku = {record["amazonAsin"]: sku for sku, record in economics.items() if record.get("amazonAsin")}

    def resolve_economics(sku: str) -> tuple[dict[str, Any], str | None]:
        if sku in economics:
            return economics[sku], sku
        variant_match = re.fullmatch(r"(.+)_(?:FBA|FAB|PSI|PSW|\d+)", sku, flags=re.IGNORECASE)
        if variant_match and variant_match.group(1) in economics:
            source_sku = variant_match.group(1)
            return economics[source_sku], source_sku
        return {}, None

    master_products: dict[str, dict[str, Any]] = {}
    ean_counts: dict[str, int] = defaultdict(int)
    for row in master_rows:
        ean = clean_id(row.get("EAN / GTIN"))
        if ean:
            ean_counts[ean] += 1
    for row in master_rows:
        sku = sku_text(row.get("Artikelnummer"))
        if not sku:
            continue
        ean = clean_id(row.get("EAN / GTIN"))
        price = round_value(number(row.get("price"), comma_decimal=True), 2)
        purchase = round_value(number(row.get("Letzter EK"), comma_decimal=True), 2)
        delivery = round_value(number(row.get("Logistikkosten"), comma_decimal=True), 2)
        landed = round_value(number(row.get("Landed Cost"), comma_decimal=True), 2)
        econ, economics_source_sku = resolve_economics(sku)
        margin = econ.get("margin")
        if margin is None and price and purchase is not None and delivery is not None:
            net_price = price / (1 + VAT_RATE)
            margin = ratio(net_price - purchase - delivery, net_price)
        master_products[sku] = {
            "sku": sku,
            "canonicalSku": sku,
            "ean": ean or None,
            "eanAmbiguous": bool(ean and ean_counts[ean] > 1),
            "asin": econ.get("amazonAsin") or "",
            "name": econ.get("description") or f"Product {sku}",
            "supplier": repair_text(row.get("Firma / Lieferant")) or None,
            "manufacturerNumber": clean_id(row.get("Hersteller-Nr.")) or None,
            "manufacturer": repair_text(row.get("Herstellername")) or None,
            "price": price,
            "margin": round_value(margin),
            "category": econ.get("category") or infer_category(margin),
            "economicsDescription": econ.get("description"),
            "unitCosts": {
                "purchaseNet": purchase,
                "deliveryNet": delivery,
                "landedNet": landed,
                "provisionRate": PROVISION_RATE,
                "sourceSku": sku,
            },
            "active": True,
        }

    def resolve_master_sku(sku: str, asin: str) -> tuple[str | None, str | None]:
        if sku in master_products:
            return sku, "exact_sku"
        variant_match = re.fullmatch(r"(.+)_(?:FBA|FAB|PSI|PSW|\d+)", sku, flags=re.IGNORECASE)
        if variant_match and variant_match.group(1) in master_products:
            return variant_match.group(1), "sku_suffix"
        economics_sku = economics_asin_to_sku.get(asin)
        if economics_sku in master_products:
            return economics_sku, "economics_asin"
        return None, None

    active_products: dict[str, dict[str, Any]] = {}
    asin_to_sku: dict[str, str] = {}
    unmatched_amazon_listings: list[dict[str, str]] = []
    for row in product_rows:
        sku = sku_text(row.get("seller-sku"))
        asin = clean_id(row.get("asin1"))
        if not sku:
            continue
        canonical_sku, match_method = resolve_master_sku(sku, asin)
        master = master_products.get(canonical_sku or "", {})
        econ, economics_source_sku = resolve_economics(canonical_sku or sku)
        if not canonical_sku:
            unmatched_amazon_listings.append({"sellerSku": sku, "asin": asin})
        unit_costs = master.get("unitCosts")
        if not unit_costs and economics_source_sku:
            unit_costs = {
                "purchaseNet": econ.get("purchaseCostNet"),
                "deliveryNet": econ.get("deliveryCostNet"),
                "landedNet": None,
                "provisionRate": PROVISION_RATE,
                "sourceSku": economics_source_sku,
            }
        active_products[sku] = {
            "sku": sku,
            "canonicalSku": canonical_sku,
            "sourceMatch": match_method,
            "ean": master.get("ean"),
            "eanAmbiguous": master.get("eanAmbiguous", False),
            "asin": asin,
            "name": repair_text(row.get("item-name")) or master.get("name") or "Unnamed product",
            "price": master.get("price") if master.get("price") is not None else round_value(number(row.get("price"), comma_decimal=True), 2),
            "active": True,
            "margin": econ.get("margin") if econ.get("margin") is not None else master.get("margin"),
            "category": econ.get("category") or master.get("category"),
            "economicsDescription": econ.get("description") or master.get("economicsDescription"),
            "unitCosts": unit_costs,
        }
        if asin:
            asin_to_sku[asin] = sku

    advertised_filtered = [
        row for row in advertised_rows
        if row.get("Budget currency") == "EUR" and row.get("Advertised product marketplace") == "AMAZON_DE"
    ]
    daily: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    product_ads: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    ad_group_products: dict[tuple[str, str], dict[str, dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    ad_dates: list[date] = []
    for row in advertised_filtered:
        current_date = parse_date(row.get("Date"))
        if not current_date:
            continue
        ad_dates.append(current_date)
        date_key = current_date.isoformat()
        sku = sku_text(row.get("Advertised product SKU"))
        asin = clean_id(row.get("Advertised product ID"))
        product_key = sku or asin
        values = {
            "impressions": metric(number(row.get("Impressions"))),
            "clicks": metric(number(row.get("Clicks"))),
            "spend": metric(number(row.get("Total cost"))),
            "purchases": metric(number(row.get("Purchases"))),
            "sales": metric(number(row.get("Sales"))),
            "units": metric(number(row.get("Units sold"))),
        }
        for key, value in values.items():
            daily[date_key][key] += value
            if product_key:
                product_ads[product_key][key] += value
        if product_key:
            product_ads[product_key]["asin"] = asin
            product_ads[product_key]["sku"] = sku
            product_ads[product_key]["name"] = repair_text(row.get("Advertised product name"))
        campaign_id = clean_id(row.get("Campaign ID"))
        ad_group_id = clean_id(row.get("Ad group ID"))
        if campaign_id and ad_group_id and product_key:
            member = ad_group_products[(campaign_id, ad_group_id)][product_key]
            member["impressions"] += values["impressions"]
            member["clicks"] += values["clicks"]
            member["spend"] += values["spend"]
            member["sales"] += values["sales"]

    ad_group_product_map: dict[tuple[str, str], dict[str, Any]] = {}
    for key, members in ad_group_products.items():
        ranked = sorted(
            members.items(),
            key=lambda item: (item[1]["spend"], item[1]["clicks"], item[1]["impressions"]),
            reverse=True,
        )
        product_key = ranked[0][0]
        ad_group_product_map[key] = {
            "productKey": product_key,
            "ambiguous": len(ranked) > 1,
            "candidateCount": len(ranked),
        }

    retail_by_sku: dict[str, dict[str, Any]] = {}
    for row in business_rows:
        sku = sku_text(row.get("SKU")) or asin_to_sku.get(clean_id(row.get("(Child) ASIN")), "")
        if not sku:
            continue
        sessions = metric(number(row.get("Sessions – Total")))
        units = metric(number(row.get("Units ordered")))
        retail_by_sku[sku] = {
            "sessions": round_value(sessions, 0),
            "pageViews": round_value(metric(number(row.get("Page views – Total"))), 0),
            "units": round_value(units, 0),
            "sales": round_value(metric(number(row.get("Ordered Product Sales"))), 2),
            "conversion": round_value(number(row.get("Unit Session Percentage"), percent=True)),
            "buyBox": round_value(number(row.get("Featured Offer (Buy Box) percentage"), percent=True)),
        }

    normalized_products = []
    advertised_active_skus = set()
    for sku, product in active_products.items():
        ad = product_ads.get(sku)
        if not ad and product.get("asin"):
            ad = product_ads.get(product["asin"])
        if ad:
            advertised_active_skus.add(sku)
        retail = retail_by_sku.get(sku)
        normalized = {
            **product,
            "retail": retail,
            "advertising": None,
            "advertisingStatus": "Observed activity" if ad else "No activity observed",
        }
        if ad:
            normalized["advertising"] = enrich_metrics({
                key: metric(ad.get(key)) for key in ("impressions", "clicks", "spend", "purchases", "sales", "units")
            })
        normalized_products.append(normalized)

    normalized_products.sort(key=lambda item: metric((item.get("advertising") or {}).get("sales")), reverse=True)

    placements: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in placement_rows:
        if row.get("Budget currency") != "EUR":
            continue
        name = repair_text(row.get("Placement classification")) or repair_text(row.get("Placement name")) or "Unclassified"
        bucket = placements[name]
        for source, target in (("Impressions", "impressions"), ("Clicks", "clicks"), ("Total cost", "spend"), ("Purchases", "purchases"), ("Sales", "sales")):
            bucket[target] += metric(number(row.get(source)))
    placement_list = [enrich_metrics({"name": name, **values}) for name, values in placements.items()]
    placement_list.sort(key=lambda item: item["spend"], reverse=True)

    campaigns = []
    for row in campaign_rows:
        if repair_text(row.get("Country")) != "Germany":
            continue
        spend = metric(number(row.get("Total cost")))
        sales = metric(number(row.get("Sales")))
        campaigns.append({
            "name": repair_text(row.get("Campaign name")),
            "state": repair_text(row.get("State")),
            "type": repair_text(row.get("Type")),
            "targeting": repair_text(row.get("Targeting")),
            "strategy": repair_text(row.get("Campaign bid strategy")),
            "budget": round_value(number(row.get("Campaign budget amount")), 2),
            "topOfSearchShare": round_value(number(row.get("Top-of-search impression share (IS)"), percent=True)),
            "topOfSearchAdjustment": round_value(number(row.get("Top-of-search bid adjustment"), percent=True)),
            "spend": round_value(spend, 2),
            "sales": round_value(sales, 2),
            "acos": round_value(ratio(spend, sales)),
            "roas": round_value(ratio(sales, spend)),
        })
    campaigns.sort(key=lambda item: item["spend"], reverse=True)

    targets: dict[str, dict[str, Any]] = {}
    target_terms: dict[str, dict[str, dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    target_dates: dict[str, list[date]] = defaultdict(list)
    for row in targeting_rows:
        if row.get("Budget currency") != "EUR":
            continue
        target_id = clean_id(row.get("Target ID"))
        campaign_id = clean_id(row.get("Campaign ID"))
        ad_group_id = clean_id(row.get("Ad group ID"))
        if not target_id:
            target_id = "|".join((campaign_id, ad_group_id, repair_text(row.get("Targeting"))))
        if target_id not in targets:
            targets[target_id] = {
                "id": target_id,
                "campaignId": campaign_id,
                "campaignName": repair_text(row.get("Campaign name")),
                "adGroupId": ad_group_id,
                "adGroupName": repair_text(row.get("Ad group name")),
                "target": repair_text(row.get("Targeting")),
                "matchType": repair_text(row.get("Targeting match type")),
                "targetType": repair_text(row.get("Target type")),
                "status": repair_text(row.get("Target status")),
                "bid": round_value(number(row.get("Target bid")), 2),
                "impressions": 0.0,
                "clicks": 0.0,
                "spend": 0.0,
                "purchases": 0.0,
                "sales": 0.0,
                "units": 0.0,
            }
        target = targets[target_id]
        for source, destination in (("Impressions", "impressions"), ("Clicks", "clicks"), ("Total cost", "spend"), ("Purchases", "purchases"), ("Sales", "sales"), ("Units sold", "units")):
            target[destination] += metric(number(row.get(source)))
        current_date = parse_date(row.get("Date"))
        if current_date:
            target_dates[target_id].append(current_date)
        term = repair_text(row.get("Search term"))
        if term:
            term_bucket = target_terms[target_id][term]
            for source, destination in (("Impressions", "impressions"), ("Clicks", "clicks"), ("Total cost", "spend"), ("Purchases", "purchases"), ("Sales", "sales")):
                term_bucket[destination] += metric(number(row.get(source)))

    target_list = []
    matched_target_count = 0
    ambiguous_target_count = 0
    for target_id, target in targets.items():
        mapping = ad_group_product_map.get((target["campaignId"], target["adGroupId"]))
        product = None
        if mapping:
            key = mapping["productKey"]
            sku = sku_text(product_ads.get(key, {}).get("sku"))
            asin = clean_id(product_ads.get(key, {}).get("asin"))
            if not sku and asin:
                sku = asin_to_sku.get(asin, "")
            product = active_products.get(sku)
            matched_target_count += int(bool(product))
            ambiguous_target_count += int(mapping["ambiguous"])
        else:
            sku = ""
            asin = ""
        dates = target_dates.get(target_id, [])
        terms = []
        for term, values in target_terms[target_id].items():
            term_record = enrich_metrics({"term": term, **values})
            terms.append(term_record)
        terms.sort(key=lambda item: (item["sales"], item["spend"], item["clicks"]), reverse=True)
        target.update({
            "sku": sku or None,
            "asin": (product or {}).get("asin") if product else (asin or None),
            "productName": (product or {}).get("name") if product else None,
            "price": (product or {}).get("price") if product else None,
            "margin": (product or {}).get("margin") if product else None,
            "category": (product or {}).get("category") if product else None,
            "ambiguousProduct": bool(mapping and mapping["ambiguous"]),
            "productCandidateCount": mapping["candidateCount"] if mapping else 0,
            "dateStart": min(dates).isoformat() if dates else None,
            "dateEnd": max(dates).isoformat() if dates else None,
            "topSearchTerms": terms[:5],
        })
        target_list.append(enrich_metrics(target))
    target_list.sort(key=lambda item: (item["spend"], item["sales"]), reverse=True)

    promotion_candidates = []
    for product in normalized_products:
        if product["advertising"] is not None:
            continue
        retail = product.get("retail") or {}
        margin = product.get("margin")
        if product.get("category") in {"Abverkauf", "Preiseinstieg"}:
            evidence = "Do not promote"
        elif margin is not None and retail.get("units", 0) > 0:
            evidence = "Strong candidate"
        elif margin is not None:
            evidence = "Potential candidate"
        else:
            evidence = "Insufficient evidence"
        score = metric(margin) * 100 + metric(retail.get("conversion")) * 100 + min(metric(retail.get("sessions")) / 100, 20)
        promotion_candidates.append({
            "sku": product["sku"],
            "asin": product["asin"],
            "name": product["name"],
            "price": product["price"],
            "margin": margin,
            "category": product.get("category"),
            "retail": product.get("retail"),
            "level": evidence,
            "score": round_value(score, 1),
            "reason": "No advertising activity was observed in the supplied 30-day Advertised Product export.",
        })
    promotion_candidates.sort(key=lambda item: item["score"], reverse=True)

    daily_list = [enrich_metrics({"date": date_key, **values}) for date_key, values in sorted(daily.items())]
    ad_totals = enrich_metrics({
        key: sum(metric(day.get(key)) for day in daily_list)
        for key in ("impressions", "clicks", "spend", "purchases", "sales", "units")
    })
    retail_totals = {
        "sales": round_value(sum(metric(value.get("sales")) for value in retail_by_sku.values()), 2),
        "sessions": round_value(sum(metric(value.get("sessions")) for value in retail_by_sku.values()), 0),
        "units": round_value(sum(metric(value.get("units")) for value in retail_by_sku.values()), 0),
    }
    retail_totals["conversion"] = round_value(ratio(retail_totals["units"], retail_totals["sessions"]))

    covered_gross_sales = 0.0
    covered_net_sales = 0.0
    purchase_cost = 0.0
    delivery_cost = 0.0
    provision_cost = 0.0
    covered_cost_products = 0
    missing_cost_products = []
    for product in normalized_products:
        retail = product.get("retail") or {}
        gross_sales = metric(retail.get("sales"))
        if gross_sales <= 0:
            continue
        unit_costs = product.get("unitCosts") or {}
        purchase_unit = unit_costs.get("purchaseNet")
        delivery_unit = unit_costs.get("deliveryNet")
        if purchase_unit is None or delivery_unit is None:
            missing_cost_products.append({"sku": product["sku"], "grossSales": round_value(gross_sales, 2)})
            continue
        units = metric(retail.get("units"))
        net_sales = gross_sales / (1 + VAT_RATE)
        covered_gross_sales += gross_sales
        covered_net_sales += net_sales
        purchase_cost += units * metric(purchase_unit)
        delivery_cost += units * metric(delivery_unit)
        provision_cost += net_sales * PROVISION_RATE
        covered_cost_products += 1

    total_cost = purchase_cost + delivery_cost + provision_cost + ad_totals["spend"]
    net_contribution = covered_net_sales - total_cost
    profitability = {
        "tcos": round_value(ratio(ad_totals["spend"], retail_totals["sales"])),
        "netContribution": round_value(net_contribution, 2),
        "netContributionMargin": round_value(ratio(net_contribution, covered_net_sales)),
        "coveredGrossSales": round_value(covered_gross_sales, 2),
        "coveredNetSales": round_value(covered_net_sales, 2),
        "retailSalesCoverage": round_value(ratio(covered_gross_sales, retail_totals["sales"])),
        "purchaseCost": round_value(purchase_cost, 2),
        "deliveryCost": round_value(delivery_cost, 2),
        "provisionCost": round_value(provision_cost, 2),
        "advertisingCost": round_value(ad_totals["spend"], 2),
        "totalCost": round_value(total_cost, 2),
        "vatRate": VAT_RATE,
        "provisionRate": PROVISION_RATE,
        "missingCostProducts": missing_cost_products,
    }

    role_meta = {
        "advertised_product": ("Advertised Product", "Dashboard source", "Ready"),
        "campaign": ("Campaign", "Budgets and campaign settings", "Ready"),
        "placement": ("Placement", "Placement analysis", "Ready"),
        "search_term_summary": ("Search Term summary", "Validation only; overlapping grain", "Validation only"),
        "search_term_daily": ("Search Term daily", "Validation only; overlaps Targeting", "Validation only"),
        "targeting": ("Targeting", "Bidding evidence", "Ready"),
        "business_report": ("Business Report", "Retail performance; partial catalog coverage", "Partial"),
        "product_master": ("Completed product master", "Canonical SKU, EAN, supplier, cost and price source of truth", "Ready"),
        "amazon_listing": ("Amazon Product List", "Amazon listing names, seller SKUs and ASINs", "Partial"),
        "economics": ("Calculation workbook", "Contribution margin and category", "Partial"),
    }
    row_counts = {
        "advertised_product": len(advertised_rows), "campaign": len(campaign_rows), "placement": len(placement_rows),
        "search_term_summary": len(search_summary_rows), "search_term_daily": len(search_daily_rows),
        "targeting": len(targeting_rows), "business_report": len(business_rows), "product_master": len(master_rows),
        "amazon_listing": len(product_rows),
        "economics": len(economics_rows),
    }
    imports = []
    manifest = []
    for key, file_path in files.items():
        relative_path = Path(os.path.relpath(file_path, source_root)).as_posix()
        digest = hash_file(file_path)
        label, role, status = role_meta[key]
        imports.append({
            "key": key,
            "file": file_path.name,
            "path": relative_path,
            "report": label,
            "role": role,
            "rows": row_counts[key],
            "status": status,
            "sha256": digest,
        })
        manifest.append({"path": relative_path, "bytes": file_path.stat().st_size, "sha256": digest})

    date_start = min(ad_dates).isoformat() if ad_dates else None
    date_end = max(ad_dates).isoformat() if ad_dates else None
    economics_coverage = sum(1 for product in normalized_products if product.get("margin") is not None)
    ad_product_keys = {sku_text(value.get("sku")) or clean_id(value.get("asin")) for value in product_ads.values()}
    unmatched_ad_product_count = sum(1 for key in ad_product_keys if key and key not in active_products and key not in asin_to_sku)

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "reporting": {
            "start": date_start,
            "end": date_end,
            "days": len(daily_list),
            "currency": "EUR",
            "marketplace": "Amazon DE",
            "timezone": "Europe/Berlin",
        },
        "settings": {
            "aggressivenessFactor": 0.70,
            "maxBidChange": 0.20,
            "minimumClicks": 5,
            "manualApproval": True,
            "evaluationCadence": "Weekly",
            "evidenceWindowDays": 30,
        },
        "totals": {"advertising": ad_totals, "retail": retail_totals, "profitability": profitability},
        "daily": daily_list,
        "placements": placement_list,
        "campaigns": campaigns,
        "products": normalized_products,
        "catalogProducts": list(master_products.values()),
        "targetPerformance": target_list,
        "promotionCandidates": promotion_candidates,
        "imports": imports,
        "quality": {
            "activeProducts": len(normalized_products),
            "masterCatalogProducts": len(master_products),
            "masterCatalogEanProducts": sum(1 for product in master_products.values() if product.get("ean")),
            "ambiguousMasterEans": sum(1 for count in ean_counts.values() if count > 1),
            "unmatchedAmazonListings": len(unmatched_amazon_listings),
            "unmatchedAmazonListingRows": unmatched_amazon_listings,
            "retailCoverageProducts": len(retail_by_sku),
            "economicsCoverageProducts": economics_coverage,
            "netContributionCoverageProducts": covered_cost_products,
            "advertisedActiveProducts": len(advertised_active_skus),
            "targets": len(target_list),
            "targetsMatchedToActiveProduct": matched_target_count,
            "ambiguousTargetProductJoins": ambiguous_target_count,
            "unmatchedAdvertisingProducts": unmatched_ad_product_count,
            "excludedNonEuroAdvertisedRows": len(advertised_rows) - len(advertised_filtered),
            "duplicateProtection": "Dashboard totals use Advertised Product only. Targeting drives recommendations; both Search Term exports are validation-only.",
        },
        "sourceManifest": manifest,
    }


def verify_sources(source_root: Path) -> None:
    if not OUTPUT_PATH.exists():
        raise FileNotFoundError(f"No normalized snapshot exists at {OUTPUT_PATH}")
    snapshot = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
    failures = []
    for item in snapshot.get("sourceManifest", []):
        file_path = source_root / item["path"]
        current = hash_file(file_path) if file_path.exists() else "missing"
        if current != item["sha256"]:
            failures.append(item["path"])
    if failures:
        raise RuntimeError("Source immutability check failed:\n" + "\n".join(failures))
    print(f"Verified {len(snapshot.get('sourceManifest', []))} immutable source files.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize immutable Amazon exports into an application snapshot.")
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--product-master", type=Path, help="Override the immutable completed product-master workbook.")
    parser.add_argument("--check", action="store_true", help="Verify current source hashes against the generated manifest.")
    args = parser.parse_args()
    source_root = args.source_root.resolve()
    if args.check:
        verify_sources(source_root)
        return
    snapshot = build_snapshot(source_root, args.product_master.resolve() if args.product_master else None)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUTPUT_PATH),
        "products": len(snapshot["products"]),
        "targets": len(snapshot["targetPerformance"]),
        "reporting": snapshot["reporting"],
        "quality": snapshot["quality"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
